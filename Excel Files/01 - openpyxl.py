# Excel documents basics

# pip install openpyxl

import openpyxl

# Opens the workbook file (xlsx)
wb = openpyxl.load_workbook('exmp.xlsx')

# Prints all sheett names
print(wb.sheetnames)

# Prints the current active sheet name
print(wb.active)

# Select a sheet by it's name
wantSheet = wb['LongBoringSheet']

print(wantSheet)

workSheet = wb.active

# Get the value of a specific cell
print(workSheet['B4'].value)

place = workSheet['B5']

print('Cell ' + place.coordinate + ' has value: ' + place.value)

# All values in the B column

for i in range(2,8):
    print(workSheet.cell(row=i, column=2).value)


# returns the highest amount of rows in any column
print(workSheet.max_row)

# returns the highest amount of columns in any row
print(workSheet.max_column)

# Getting rows and columns

for row in workSheet['A1':'C6']:
    print('===ROW START===')
    for obj in row:
        print(obj.coordinate, obj.value)
    print('===ROW END===')


