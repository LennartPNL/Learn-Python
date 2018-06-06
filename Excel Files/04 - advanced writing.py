# More features while writing to excel files

import openpyxl
from openpyxl.styles import *

wb = openpyxl.Workbook()

sheet = wb['Sheet']

print(sheet.title)

# Attributes of your font
myFont = Font(size=32, bold=True)

# making your own style
myStyle = NamedStyle('myStyle')

# Adding your font to your style
myStyle.font = myFont

# Adding the style to a cell
sheet['A1'].style = myStyle

# Writing text to the cell
sheet['A1'] = 'Look at my font'

# Saving the file
wb.save('withStyle.xlsx')

# Another example #

myFont2 = Font('Comic Sans MS')

myStyle2 = NamedStyle('myStyle2')

myStyle2.font = myFont2

sheet['A2'].style = myStyle2

sheet['A2'] = 'Please do not use Comic Sans'

wb.save('withStyle.xlsx')



# Formulas #
from random import *
sheet['G1'] = randint(1, 9999)
sheet['G2'] = randint(1, 9999)
sheet['G3'] = randint(1, 9999)
sheet['G4'] = randint(1, 9999)
sheet['G5'] = randint(1, 9999)
sheet['G6'] = '=SUM(G1:G5)'

wb.save('withStyle.xlsx')

# Column Size #

myHeight = randint(20, 100)
myWidth = randint(20, 100)

sheet.row_dimensions[10].height = myHeight
sheet.column_dimensions['H'].width = myWidth

sheet['A10'] = 'This row is %s pixels high' % str(myHeight)
sheet['H1'] = 'This column is %s pixels wide' % str(myWidth)

wb.save('withStyle.xlsx')


# Merging Cells #

sheet.merge_cells('I1:K5')
sheet.merge_cells('L10:P12')

sheet['I1'] = 'All these cells are merged'

sheet.unmerge_cells('L10:P12')

sheet['L10'] = 'There used to be some merged cells here'

wb.save('withStyle.xlsx')

# Freezing Panes #

# Freezes all cells up to A2, not including A2
sheet.freeze_panes = 'A2'

sheet['M1'] = 'The First Row Is A Frozen Pane'

wb.save('withStyle.xlsx')


# Charts #

from openpyxl.chart import *

sheet['A14'] = 'John'
sheet['B14'] = 'Eyong Tarkang Enoh'

# Generating some dummy data
for i in range(15, 26):
    sheet['A' + str(i)] = randint(0, 999)
    sheet['B' + str(i)] = randint(0, 999)
    sheet['C' + str(i)] = 'abcdefghijk'[i-15].upper()


chart1 = BarChart()
# Setting the chart's type, style and other visual things
chart1.type = "col"
chart1.style = randint(0, 15)
chart1.title = "Bar Chart"
chart1.y_axis.title = 'Sales'
chart1.x_axis.title = 'Period'

# Defining columns and rows from which values have to be in the chart
data = Reference(sheet, min_col=1, min_row=14, max_row=25, max_col=2)

# Setting categories
cats = Reference(sheet, min_col=3, min_row=15, max_row=25)

# Adding the data with titles
chart1.add_data(data, titles_from_data=True)

# Setting the categories
chart1.set_categories(cats)
chart1.shape = 4

# From where the chart will start (top left corner)
sheet.add_chart(chart1, "F17")

# Adding some totals
sheet['A26'] = '=SUM(A15:A25)'
sheet['B26'] = '=SUM(B15:B25)'

sheet['A26'].font = Font(bold=True)
sheet['B26'].font = Font(bold=True)

# Readable width
sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 20

wb.save('withStyle.xlsx')




