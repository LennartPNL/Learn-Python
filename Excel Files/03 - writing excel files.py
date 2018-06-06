# Writing to excel files

import openpyxl

wb = openpyxl.Workbook()

print(wb.sheetnames)

sheet = wb.active

print(sheet.title)


# Changing the name of the sheet
sheet.title = 'I just took a sheet'

print(sheet.title)

# Saving the workbook to a file
wb.save('changedSheet.xlsx')

### New workbook ###

workbook = openpyxl.Workbook()

# Adds a new sheet to the file
workbook.create_sheet()

print(workbook.sheetnames)

workbook.create_sheet(index=0, title='First Sheet')

workbook.create_sheet(index=2, title='Cool Sheet')

print(workbook.sheetnames)

# Removing sheets

workbook.remove(workbook['Sheet1'])

print(workbook.sheetnames)

# Writing values

worksheet = workbook['First Sheet']

worksheet['A1'] = 'Title of column A'

print(worksheet['A1'].value)

workbook.save('writecells.xlsx')